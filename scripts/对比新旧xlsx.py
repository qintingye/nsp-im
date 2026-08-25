#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""对比新旧版 xlsx 题库 ·安全索引"""
import openpyxl
from collections import Counter
from pathlib import Path

OLD = r'C:\Users\Administrator\AppData\Local\hermes\cache\documents\doc_871ecd55441d_新型电力系统岗位题库-更新.xlsx'
NEW = r'C:\Users\Administrator\AppData\Local\hermes\cache\documents\doc_354b3f59dd91_新型电力系统岗位题库-更新(1).xlsx'

def safe_get(row, idx):
    return row[idx] if 0 <= idx < len(row) else None

def extract(xlsx):
    wb = openpyxl.load_workbook(xlsx, data_only=True)
    qs = []
    for s in wb.sheetnames:
        ws = wb[s]
        headers = [c.value for c in ws[1]]
        IDX = {
            'id': next((i for i, h in enumerate(headers) if h == '序号'), 0),
            'cat': next((i for i, h in enumerate(headers) if h in ('制度名称', '题库名称', '题库分类')), 2),
            'type': next((i for i, h in enumerate(headers) if h == '题型'), 3),
            'diff': next((i for i, h in enumerate(headers) if h == '试题难度'), 4),
            'q': next((i for i, h in enumerate(headers) if h == '试题题目'), 8),
            'opts_start': next((i for i, h in enumerate(headers) if h and '选项A' in str(h)), 9),
            'ans': next((i for i, h in enumerate(headers) if h == '答案'), 17),
            'exp': next((i for i, h in enumerate(headers) if h == '题目解析'), 18),
        }
        for row in ws.iter_rows(min_row=2, values_only=True):
            rid = safe_get(row, IDX['id'])
            if not rid: continue
            opts = []
            for j in range(8):
                v = safe_get(row, IDX['opts_start'] + j)
                if v: opts.append(v)
            ans = safe_get(row, IDX['ans'])
            qs.append({
                'id': rid,
                'cat': safe_get(row, IDX['cat']) or '',
                'type': safe_get(row, IDX['type']) or '单选题',
                'diff': safe_get(row, IDX['diff']) or '简单',
                'q': (safe_get(row, IDX['q']) or '').replace('\r\n', ' ').replace('\n', ' ').strip(),
                'opts': opts,
                'answer': ans if isinstance(ans, str) else (''.join(ans) if ans else ''),
                'has_answer': bool(ans),
                'explain': (safe_get(row, IDX['exp']) or '暂无解析').strip(),
            })
    return qs

old_q = extract(OLD)
new_q = extract(NEW)

print(f'=== 旧版 ===')
print(f'  总题数: {len(old_q)}, 有答案: {sum(1 for q in old_q if q["has_answer"])}')

print(f'\n=== 新版 ===')
print(f'  总题数: {len(new_q)}, 有答案: {sum(1 for q in new_q if q["has_answer"])}')
print(f'  题型: {dict(Counter(q["type"] for q in new_q))}')

old_ids = {q['id'] for q in old_q}
new_ids = {q['id'] for q in new_q}
added = new_ids - old_ids
removed = old_ids - new_ids
print(f'\n=== 差异 ===')
print(f'  新增 ID: {len(added)} 个')
if added:
    print(f'    示例: {sorted(added)[:15]}')
print(f'  删除 ID: {len(removed)} 个')
if removed:
    print(f'    示例: {sorted(removed)[:15]}')

old_map = {q['id']: q for q in old_q}
new_ans_added = 0
ans_changed = 0
for q in new_q:
    if q['id'] in old_map:
        oq = old_map[q['id']]
        if q['has_answer'] and not oq['has_answer']:
            new_ans_added += 1
        elif q['has_answer'] and oq['has_answer'] and q['answer'] != oq['answer']:
            ans_changed += 1

print(f'\n=== 答案变化 ===')
print(f'  之前没答案现在有: {new_ans_added} 题')
print(f'  答案修改: {ans_changed} 题')

# 保存新版 JSON
import json
out = Path(r'C:/Users/Administrator/AppData/Local/Temp/all_questions_v4.json')
with open(out, 'w', encoding='utf-8') as f:
    json.dump(new_q, f, ensure_ascii=False, indent=2)
print(f'\n已保存新版 {len(new_q)} 题到 {out.name}')