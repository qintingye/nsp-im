#!/usr/bin/env python3
"""
打包"新型电力系统岗位题库"为 ZIP
- 包含 HTML 主程序
- 包含 questions.json 题库数据（80 题）
- 输出到桌面，方便传输到手机
"""
import os
import json
import zipfile
import shutil
from pathlib import Path

OUTPUT_DIR = Path(r'C:\Users\Administrator\Desktop')
PKG_NAME = '新型电力系统岗位题库_v1.0'
OUTPUT_ZIP = OUTPUT_DIR / f'{PKG_NAME}.zip'

HTML_PATH = Path(r'D:\Obsidian-Knowledge\01-Domain\新型电力系统建设\学习成果\新型电力系统岗位题库.html')
XLSX_PATH = Path(r'C:\Users\Administrator\AppData\Local\hermes\cache\documents\doc_3cd56d6150fe_新型电力系统岗位题库.xlsx')
QUESTIONS_JSON = Path(r'D:\Obsidian-Knowledge\01-Domain\新型电力系统建设\学习成果\questions.json')

print('=' * 60)
print(f'📦 打包工具 · {PKG_NAME}')
print('=' * 60)

# 1. 把 xlsx 转成 questions.json
import openpyxl

print('\n[1/4] 读取 xlsx 题库...')
wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)

all_questions = []
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    headers = [c.value for c in ws[1]]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        q = {
            'id': row[0],
            'cat': row[1],
            'type': row[2],
            'difficulty': row[3],
            'is_secret': row[4],
            'keywords': row[5],
            'qid': row[6],
            'q': row[7],
            'opts': [opt for opt in row[8:14] if opt],
            'answer': row[15] if isinstance(row[15], str) else (','.join(row[15]) if row[15] else ''),
            'explain': row[16] or '暂无解析（公司内部资料未提供）',
            'public': row[17]
        }
        all_questions.append(q)
print(f'    共读取 {len(all_questions)} 道题')

# 2. 写 questions.json
print('\n[2/4] 生成 questions.json...')
out = {
    'meta': {
        'name': '新型电力系统岗位题库',
        'version': '1.0',
        'total': len(all_questions),
        'source': '公司内部资料 · 仅限本人学习使用',
        'public': False,
        'warning': '本卷标记为"是否公开=否"，请勿外传'
    },
    'questions': all_questions
}
with open(QUESTIONS_JSON, 'w', encoding='utf-8') as f:
    json.dump(out, f, ensure_ascii=False, indent=2)
print(f'    已生成 questions.json ({len(json.dumps(out))} 字符)')

# 3. 验证 HTML 文件
print('\n[3/4] 检查 HTML 主程序...')
if not HTML_PATH.exists():
    print(f'    ❌ 错误：HTML 文件不存在 {HTML_PATH}')
    exit(1)
print(f'    HTML 大小: {HTML_PATH.stat().st_size / 1024:.1f} KB')

# 4. 打 ZIP
print(f'\n[4/4] 打包 ZIP 到 {OUTPUT_ZIP}...')
if OUTPUT_ZIP.exists():
    OUTPUT_ZIP.unlink()

with zipfile.ZipFile(OUTPUT_ZIP, 'w', zipfile.ZIP_DEFLATED, compresslevel=6) as zf:
    zf.write(HTML_PATH, arcname=f'{PKG_NAME}/新型电力系统岗位题库.html')
    zf.write(QUESTIONS_JSON, arcname=f'{PKG_NAME}/questions.json')
    # 加一个 README
    readme = f'''# 新型电力系统岗位题库 v1.0

## 📦 内容
- 新型电力系统岗位题库.html - 主程序（双击即可在浏览器打开）
- questions.json - 题库数据（{len(all_questions)} 道题）

## 📱 使用方法
### 电脑端
1. 解压 ZIP
2. 双击 `新型电力系统岗位题库.html` 即可使用

### 手机端
1. 把整个文件夹传到手机（微信文件/QQ/云盘/数据线均可）
2. 用手机浏览器打开 HTML 文件
3. iOS Safari / Android Chrome 可"添加到主屏"，像 App 一样使用

## 🔒 隐私声明
- 本资料标记为"是否公开=否"
- 仅限本人学习使用，请勿外传
- 解析为官方文档原文，已尽量准确，但以官方文件为准

## 📊 题库分布
- 总题数: {len(all_questions)} 道
- 6 大制度: 公司并网服务管理办法 / 电力基建项目前期管理办法 / 电网规划管理办法 / 新能源管理细则 / 投资管理规定 / 战略管理规定
- 题型: 单选题 / 多选题 / 不定项选择题

## 🐛 反馈
如发现题目或答案错误，请用记事本修改 questions.json 即可
'''
    zf.writestr(f'{PKG_NAME}/README.md', readme)

size_kb = OUTPUT_ZIP.stat().st_size / 1024
print(f'    ✅ 打包完成: {OUTPUT_ZIP}')
print(f'    大小: {size_kb:.1f} KB')

print('\n' + '=' * 60)
print('🎉 完成！')
print('=' * 60)
print(f'\n📦 文件位置: {OUTPUT_ZIP}')
print(f'📏 大小: {size_kb:.1f} KB')
print('\n🚀 下一步：')
print('1. 用微信/QQ/邮箱把 ZIP 发到手机')
print('2. 手机用浏览器打开 HTML')
print('3. iOS Safari 添加到主屏（像 App 一样用）')
print('4. Android Chrome 同样支持')