#!/usr/bin/env python3
"""闯关版 v2：答错只回上一关（不是重头开始）"""
import openpyxl, json
from pathlib import Path

XLSX = Path(r'C:\Users\Administrator\AppData\Local\hermes\cache\documents\doc_3cd56d6150fe_新型电力系统岗位题库.xlsx')
OUT_HTML = Path(r'D:\新型电力系统题库\新型电力系统岗位题库-闯关版.html')

wb = openpyxl.load_workbook(XLSX, data_only=True)
all_questions = []
COL_ID, COL_CAT, COL_TYPE, COL_DIFF, COL_Q = 0, 2, 3, 4, 8
COL_OPTS_START, COL_OPTS_END = 9, 16
COL_ANSWER, COL_EXPLAIN = 17, 18

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    headers = [c.value for c in ws[1]]
    if headers[COL_TYPE] != '题型':
        COL_CAT_MAP = {h: i for i, h in enumerate(headers)}
        COL_TYPE = COL_CAT_MAP.get('题型', 3)
        COL_CAT = COL_CAT_MAP.get('题库分类', COL_CAT_MAP.get('制度名称', 2))
        COL_DIFF = COL_CAT_MAP.get('试题难度', 4)
        COL_Q = COL_CAT_MAP.get('试题题目', 8)
        COL_ANSWER = COL_CAT_MAP.get('答案', 17)
        COL_EXPLAIN = COL_CAT_MAP.get('题目解析', 18)
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[COL_ID]: continue
        ans = row[COL_ANSWER] if len(row) > COL_ANSWER else None
        if not ans: continue
        ans_str = ans if isinstance(ans, str) else ''.join(ans)
        all_questions.append({
            'id': row[COL_ID], 'cat': row[COL_CAT] or '',
            'type': row[COL_TYPE] or '单选题', 'difficulty': row[COL_DIFF] or '简单',
            'q': (row[COL_Q] or '').replace('\r\n', ' ').replace('\n', ' ').strip(),
            'opts': [opt for opt in row[COL_OPTS_START:COL_OPTS_END] if opt],
            'answer': ans_str, 'explain': (row[COL_EXPLAIN] or '暂无解析').strip(),
        })

print(f'共读取 {len(all_questions)} 道题')
questions_json = json.dumps(all_questions, ensure_ascii=False)

HTML_TEMPLATE = r'''<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1.0,maximum-scale=1.0">
<title>新型电力系统岗位题库 · 闯关版</title>
<style>
* { margin: 0; padding: 0; box-sizing: border-box; -webkit-tap-highlight-color: transparent; }
body { font-family: -apple-system, "PingFang SC", "Microsoft YaHei", sans-serif; background: #f5f5f5; color: #333; max-width: 480px; margin: 0 auto; min-height: 100vh; padding-bottom: 20px; }
.header { background: white; padding: 14px 16px; box-shadow: 0 1px 3px rgba(0,0,0,0.06); position: sticky; top: 0; z-index: 100; display: flex; align-items: center; justify-content: space-between; }
.header h1 { font-size: 18px; font-weight: 600; flex: 1; text-align: center; }
.header .badge { background: linear-gradient(135deg, #f59e0b, #ef4444); color: white; padding: 4px 10px; border-radius: 12px; font-size: 11px; font-weight: 600; }
.progress-section { background: white; padding: 12px 16px; display: flex; align-items: center; gap: 12px; border-bottom: 1px solid #f0f0f0; position: sticky; top: 51px; z-index: 99; }
.progress-bar { flex: 1; height: 8px; background: #f3f4f6; border-radius: 4px; overflow: hidden; }
.progress-fill { height: 100%; background: linear-gradient(90deg, #f59e0b, #ef4444); transition: width 0.3s; }
.progress-text { font-size: 12px; color: #6b7280; font-weight: 600; }
.level-card { background: linear-gradient(135deg, #fff7ed, #fed7aa); margin: 12px 16px; padding: 16px; border-radius: 12px; border-left: 4px solid #f59e0b; }
.level-title { font-size: 14px; color: #9a3412; font-weight: 600; margin-bottom: 4px; }
.level-subtitle { font-size: 12px; color: #c2410c; }
.quiz-section { background: white; margin: 12px 16px; border-radius: 12px; padding: 18px; box-shadow: 0 1px 3px rgba(0,0,0,0.04); }
.q-header { display: flex; justify-content: space-between; align-items: center; margin-bottom: 14px; padding-bottom: 10px; border-bottom: 1px solid #f0f0f0; }
.q-meta { font-size: 12px; color: #999; }
.q-type { background: #fef3c7; color: #92400e; padding: 3px 10px; border-radius: 4px; font-size: 12px; font-weight: 600; }
.q-type.multi { background: #dbeafe; color: #1e40af; }
.q-text { font-size: 16px; line-height: 1.7; color: #1f2937; margin-bottom: 16px; }
.options { display: flex; flex-direction: column; gap: 10px; }
.option { display: flex; align-items: flex-start; gap: 10px; padding: 14px; border: 2px solid #e5e7eb; border-radius: 8px; cursor: pointer; transition: all 0.2s; background: #fafafa; }
.option:hover:not(.locked) { border-color: #f59e0b; background: #fffbeb; transform: translateX(2px); }
.option.selected { border-color: #f59e0b; background: #fffbeb; }
.option.locked { cursor: not-allowed; opacity: 0.85; }
.option.wrong { border-color: #ef4444; background: #fef2f2; color: #991b1b; animation: shake 0.4s; }
.option.correct-show { border-color: #10b981; background: #ecfdf5; color: #065f46; }
.opt-letter { width: 28px; height: 28px; border-radius: 50%; background: white; border: 2px solid #d1d5db; display: flex; align-items: center; justify-content: center; font-size: 13px; font-weight: 700; flex-shrink: 0; }
.option.selected .opt-letter { background: #f59e0b; color: white; border-color: #f59e0b; }
.option.correct-show .opt-letter { background: #10b981; color: white; border-color: #10b981; }
.opt-text { flex: 1; font-size: 14px; line-height: 1.5; }
@keyframes shake { 0%, 100% { transform: translateX(0); } 25% { transform: translateX(-6px); } 75% { transform: translateX(6px); } }
.feedback { margin-top: 14px; padding: 14px; border-radius: 8px; font-size: 14px; line-height: 1.6; display: none; }
.feedback.show { display: block; animation: fadeIn 0.3s; }
.feedback.correct { background: #ecfdf5; border-left: 4px solid #10b981; color: #065f46; }
.feedback b { display: block; margin-bottom: 4px; font-size: 15px; }
.feedback .icon { font-size: 22px; margin-right: 6px; vertical-align: middle; }
.feedback .reason { font-size: 13px; color: #4b5563; margin-top: 8px; padding-top: 8px; border-top: 1px solid rgba(0,0,0,0.06); }
@keyframes fadeIn { from { opacity: 0; transform: translateY(-4px); } to { opacity: 1; transform: translateY(0); } }
.actions { display: flex; gap: 8px; margin-top: 16px; }
.actions button { flex: 1; padding: 14px; border: none; border-radius: 8px; font-size: 15px; font-weight: 700; cursor: pointer; transition: all 0.2s; }
.btn-submit { background: #f59e0b; color: white; }
.btn-submit:hover:not(:disabled) { background: #d97706; }
.btn-submit:disabled { background: #d1d5db; color: #9ca3af; cursor: not-allowed; }
.btn-next { background: #10b981; color: white; animation: pulse 2s infinite; }
.btn-next:disabled { background: #d1d5db; color: #9ca3af; cursor: not-allowed; animation: none; }
@keyframes pulse { 0% { box-shadow: 0 0 0 0 rgba(16, 185, 129, 0.5); } 70% { box-shadow: 0 0 0 10px rgba(16, 185, 129, 0); } 100% { box-shadow: 0 0 0 0 rgba(16, 185, 129, 0); } }
.fail-modal { position: fixed; top: 0; left: 0; right: 0; bottom: 0; background: rgba(0,0,0,0.6); display: none; align-items: center; justify-content: center; z-index: 200; }
.fail-modal.show { display: flex; animation: fadeIn 0.3s; }
.fail-box { background: white; padding: 28px; border-radius: 16px; max-width: 360px; margin: 0 20px; text-align: center; }
.fail-icon { font-size: 60px; margin-bottom: 12px; }
.fail-title { font-size: 22px; font-weight: 700; color: #ef4444; margin-bottom: 8px; }
.fail-text { font-size: 14px; color: #4b5563; margin-bottom: 20px; line-height: 1.6; white-space: pre-line; text-align: left; }
.fail-buttons { display: flex; gap: 8px; }
.fail-buttons button { flex: 1; padding: 12px; border: none; border-radius: 8px; font-size: 14px; font-weight: 600; cursor: pointer; }
.btn-back { background: #f59e0b; color: white; }
.btn-restart { background: #6b7280; color: white; }
.win-modal { position: fixed; top: 0; left: 0; right: 0; bottom: 0; background: linear-gradient(135deg, rgba(251,191,36,0.95), rgba(239,68,68,0.95)); display: none; align-items: center; justify-content: center; z-index: 200; }
.win-modal.show { display: flex; animation: fadeIn 0.5s; }
.win-box { background: white; padding: 32px; border-radius: 20px; max-width: 360px; margin: 0 20px; text-align: center; }
.win-icon { font-size: 80px; margin-bottom: 16px; animation: bounce 1s infinite; }
@keyframes bounce { 0%, 100% { transform: translateY(0); } 50% { transform: translateY(-10px); } }
.win-title { font-size: 26px; font-weight: 800; background: linear-gradient(135deg, #f59e0b, #ef4444); -webkit-background-clip: text; -webkit-text-fill-color: transparent; margin-bottom: 8px; }
.win-text { font-size: 14px; color: #4b5563; margin-bottom: 20px; line-height: 1.6; }
.win-stat { background: #f9fafb; padding: 14px; border-radius: 8px; margin-bottom: 20px; }
.win-stat div { font-size: 13px; color: #6b7280; margin: 4px 0; }
.win-stat div b { color: #1f2937; font-size: 18px; }
.win-box button { background: linear-gradient(135deg, #f59e0b, #ef4444); color: white; border: none; padding: 14px 28px; border-radius: 8px; font-size: 15px; font-weight: 700; cursor: pointer; }
.watermark { position: fixed; top: 50%; left: 50%; transform: translate(-50%, -50%) rotate(-30deg); font-size: 80px; color: rgba(0,0,0,0.04); font-weight: 900; pointer-events: none; z-index: 1; }
.privacy-badge { background: #fef3c7; border: 1px solid #fcd34d; color: #92400e; padding: 8px 12px; margin: 12px 16px; border-radius: 6px; font-size: 12px; line-height: 1.5; }
</style>
</head>
<body>
<div class="watermark">内部资料</div>
<div class="header">
<h1>🚧 闯关答题 · 第 <span id="level-num">1</span> 关</h1>
<span class="badge">🔥 必胜模式</span>
</div>
<div class="progress-section">
<span class="progress-text">已闯 <span id="passed">0</span> / <span id="total">__TOTAL__</span> 关</span>
<div class="progress-bar"><div class="progress-fill" id="progress-fill" style="width:0%"></div></div>
<span class="progress-text" id="streak-text">🔥 连胜 0</span>
</div>
<div class="privacy-badge">
🔒 <b>内部资料 · 仅限本人学习使用</b><br>
闯关模式：必须选对才能进入下一关。做错<b>只回到上一关</b>（不是从头开始）。
</div>
<div class="level-card">
<div class="level-title">🎯 当前关卡</div>
<div class="level-subtitle" id="level-desc">单选题 · 来自《公司并网服务管理办法》</div>
</div>
<div id="quiz-container"></div>

<div class="fail-modal" id="fail-modal">
<div class="fail-box">
<div class="fail-icon">❌</div>
<div class="fail-title">闯关失败</div>
<div class="fail-text" id="fail-text">很遗憾，你答错了这一题。\n正确答案已显示。\n再接再厉！</div>
<div class="fail-buttons">
<button class="btn-back" id="btn-back" onclick="goBack()">↩️ 回到上一关</button>
<button class="btn-restart" onclick="restartFromBeginning()">🔄 从头开始</button>
</div>
</div>
</div>

<div class="win-modal" id="win-modal">
<div class="win-box">
<div class="win-icon">🏆</div>
<div class="win-title">全部通关！</div>
<div class="win-text">🎉 恭喜你完成了全部题库的挑战！</div>
<div class="win-stat">
<div>已闯关数：<b><span id="win-passed">0</span></b> 关</div>
<div>最佳连胜：<b><span id="win-streak">0</span></b> 关</div>
<div>完成时间：<b><span id="win-time">0</span></b></div>
</div>
<button onclick="restartFromBeginning()">🔄 再来一次</button>
</div>
</div>

<script>
const QUESTIONS = __QUESTIONS__;
let current = 0;
let passed = 0;
let streak = 0;
let maxStreak = 0;
let locked = false;
let selectedIdxs = [];
let startTime = Date.now();
let lastPassed = 0;  // 记录"上一关已通过的关卡"

document.getElementById('total').textContent = QUESTIONS.length;

function renderQuestion() {
  const q = QUESTIONS[current];
  const isMulti = q.type === '多选题' || q.type === '不定项选择题';
  locked = false;
  selectedIdxs = [];
  document.getElementById('level-num').textContent = current + 1;
  document.getElementById('level-desc').innerHTML = `${q.type} · 来自《${q.cat}》`;
  document.getElementById('progress-fill').style.width = `${(current / QUESTIONS.length) * 100}%`;
  document.getElementById('passed').textContent = passed;
  document.getElementById('streak-text').textContent = `🔥 连胜 ${streak}`;
  document.getElementById('quiz-container').innerHTML = `
    <div class="quiz-section">
      <div class="q-header">
        <span class="q-type ${isMulti ? 'multi' : ''}">${q.type}${isMulti ? '（多选）' : ''}</span>
        <span class="q-meta">📊 ${q.difficulty}</span>
      </div>
      <div class="q-text">${q.q}</div>
      <div class="options" id="options">
        ${q.opts.map((opt, idx) => `
          <div class="option" onclick="selectOption(${idx}, ${isMulti})" data-idx="${idx}">
            <div class="opt-letter">${String.fromCharCode(65 + idx)}</div>
            <div class="opt-text">${opt}</div>
          </div>
        `).join('')}
      </div>
      <div class="feedback" id="feedback"></div>
      <div class="actions">
        ${isMulti ? `<button class="btn-submit" id="btn-submit" onclick="submitMulti()" disabled>提交答案</button>` : ''}
        <button class="btn-next" id="btn-next" onclick="nextQuestion()" disabled>${current === QUESTIONS.length - 1 ? '🏁 完成' : '下一关 →'}</button>
      </div>
    </div>
  `;
  window.scrollTo({top: 0, behavior: 'smooth'});
}

function selectOption(idx, isMulti) {
  if (locked) return;
  const opts = document.querySelectorAll('.option');
  if (!isMulti) {
    selectedIdxs = [idx];
    opts.forEach(o => { o.classList.remove('selected'); });
    opts[idx].classList.add('selected');
    judgeSingle();
  } else {
    if (selectedIdxs.includes(idx)) {
      selectedIdxs = selectedIdxs.filter(i => i !== idx);
      opts[idx].classList.remove('selected');
    } else {
      selectedIdxs.push(idx);
      opts[idx].classList.add('selected');
    }
    document.getElementById('btn-submit').disabled = selectedIdxs.length === 0;
  }
}

function judgeSingle() {
  const q = QUESTIONS[current];
  const opts = document.querySelectorAll('.option');
  const correctIdxs = q.answer.split('').map(a => a.charCodeAt(0) - 65);
  const isCorrect = selectedIdxs.length === correctIdxs.length && selectedIdxs.every(i => correctIdxs.includes(i));
  if (isCorrect) {
    handleCorrect(q, correctIdxs, opts);
  } else {
    handleWrong(q, correctIdxs, opts);
  }
}

function submitMulti() {
  const q = QUESTIONS[current];
  const opts = document.querySelectorAll('.option');
  const correctIdxs = q.answer.split('').map(a => a.charCodeAt(0) - 65);
  const correctSet = new Set(correctIdxs);
  const selectedSet = new Set(selectedIdxs);
  const isCorrect = correctSet.size === selectedSet.size && [...correctSet].every(x => selectedSet.has(x));
  if (isCorrect) {
    handleCorrect(q, correctIdxs, opts);
  } else {
    handleWrong(q, correctIdxs, opts);
  }
}

function handleCorrect(q, correctIdxs, opts) {
  locked = true;
  opts.forEach((o, idx) => {
    o.classList.add('locked');
    if (correctIdxs.includes(idx)) o.classList.add('correct-show');
  });
  passed++;
  lastPassed = current;  // 记录"最后通过的关卡"
  streak++;
  if (streak > maxStreak) maxStreak = streak;
  const fb = document.getElementById('feedback');
  fb.className = 'feedback show correct';
  fb.innerHTML = `<b><span class="icon">✅</span>闯关成功！答案：${q.answer}</b>${q.explain ? '<div class="reason">💡 ' + q.explain + '</div>' : ''}`;
  const subBtn = document.getElementById('btn-submit');
  if (subBtn) subBtn.style.display = 'none';
  document.getElementById('btn-next').disabled = false;
  document.getElementById('streak-text').textContent = `🔥 连胜 ${streak}`;
}

function handleWrong(q, correctIdxs, opts) {
  locked = true;
  opts.forEach((o, idx) => {
    o.classList.add('locked');
    if (selectedIdxs.includes(idx) && !correctIdxs.includes(idx)) {
      o.classList.add('wrong');
    }
    if (correctIdxs.includes(idx)) {
      o.classList.add('correct-show');
    }
  });
  const correctSet = new Set(correctIdxs);
  const selectedSet = new Set(selectedIdxs);
  const missing = [...correctSet].filter(x => !selectedSet.has(x));
  const wrong = [...selectedSet].filter(x => !correctSet.has(x));
  let reason = '';
  if (missing.length > 0 && wrong.length === 0) {
    reason = `❌ 少选了 ${missing.length} 个选项 (${missing.map(i => String.fromCharCode(65 + i)).join('/')})`;
  } else if (wrong.length > 0 && missing.length === 0) {
    reason = `❌ 选错了 ${wrong.length} 个选项 (${wrong.map(i => String.fromCharCode(65 + i)).join('/')})`;
  } else if (wrong.length > 0 && missing.length > 0) {
    reason = `❌ 选错了 ${wrong.length} 个 (${wrong.map(i => String.fromCharCode(65 + i)).join('/')})，少选了 ${missing.length} 个 (${missing.map(i => String.fromCharCode(65 + i)).join('/')})`;
  } else {
    reason = '❌ 答错了';
  }
  streak = 0;
  // 关键修复：答错后回到"上一关"（不是从头开始）
  // current 还没自增，所以"上一关" = current - 1
  // 第 1 关做错 → 没有更早的关，只能重答第 1 关
  let targetLevel;
  if (current === 0) {
    targetLevel = 1;  // 第 1 关 → 重答
  } else {
    targetLevel = current;  // 第 N 关 → 回到第 N 关（即 current - 1 + 1 显示）
  }
  document.getElementById('fail-text').innerHTML = `${reason}<br><br>正确答案：<b>${q.answer}</b>${q.explain ? '<br><br>💡 解析：' + q.explain : ''}<br><br>📍 闯关失败将<b>回到第 ${current} 关</b>重新挑战`;
  // 控制"回到上一关"按钮的可见性
  const btnBack = document.getElementById('btn-back');
  btnBack.textContent = current === 0 ? '🔄 重新挑战第 1 关' : `↩️ 回到第 ${current} 关`;
  setTimeout(() => {
    document.getElementById('fail-modal').classList.add('show');
  }, 800);
  document.getElementById('streak-text').textContent = `🔥 连胜 0`;
}

function nextQuestion() {
  if (current < QUESTIONS.length - 1) {
    current++;
    renderQuestion();
  } else {
    showWin();
  }
}

function goBack() {
  // 答错只回到上一关（避免大量重复）
  // current 还没自增，所以"上一关"就是 current - 1
  // 例如答错第 5 关 → current=4（0-indexed）→ 回到 current=4 重答第 5 关
  // 注意：passed 不变（已通过的关卡不丢失）
  if (current > 0) {
    current = current - 1;  // 回到上一关
  }
  // 如果 current === 0（就是第 1 关做错），保持不变，仍重答第 1 关
  document.getElementById('fail-modal').classList.remove('show');
  renderQuestion();
}

function restartFromBeginning() {
  current = 0;
  passed = 0;
  streak = 0;
  maxStreak = 0;
  startTime = Date.now();
  document.getElementById('fail-modal').classList.remove('show');
  document.getElementById('win-modal').classList.remove('show');
  renderQuestion();
}

function showWin() {
  document.getElementById('win-passed').textContent = passed;
  document.getElementById('win-streak').textContent = maxStreak;
  const sec = Math.round((Date.now() - startTime) / 1000);
  const m = Math.floor(sec / 60);
  const s = sec % 60;
  document.getElementById('win-time').textContent = `${m}分${s}秒`;
  document.getElementById('win-modal').classList.add('show');
}

renderQuestion();
</script>
</body>
</html>
'''

html = HTML_TEMPLATE.replace('__TOTAL__', str(len(all_questions))).replace('__QUESTIONS__', questions_json)

OUT_HTML.parent.mkdir(parents=True, exist_ok=True)
OUT_HTML.write_text(html, encoding='utf-8')
print(f'✅ 闯关版 v2 已生成: {OUT_HTML}')
print(f'   大小: {OUT_HTML.stat().st_size / 1024:.1f} KB')
print(f'   内嵌题数: {len(all_questions)}')