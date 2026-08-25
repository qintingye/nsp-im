#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""闯关版 v3 ·基于新版题库 (70 题 ·公司通用性制度题库)"""
import openpyxl, json
from pathlib import Path
from collections import Counter

XLSX = Path(r'C:\Users\Administrator\AppData\Local\hermes\cache\documents\doc_871ecd55441d_新型电力系统岗位题库-更新.xlsx')
OUT_HTML = Path(r'D:\新型电力系统题库\新型电力系统岗位题库-闯关版.html')

wb = openpyxl.load_workbook(XLSX, data_only=True)
all_questions = []

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    headers = [c.value for c in ws[1]]
    IDX_ID = next((i for i, h in enumerate(headers) if h == '序号'), 0)
    IDX_CAT = next((i for i, h in enumerate(headers) if h in ('制度名称', '题库名称', '题库分类')), 2)
    IDX_TYPE = next((i for i, h in enumerate(headers) if h == '题型'), 3)
    IDX_DIFF = next((i for i, h in enumerate(headers) if h == '试题难度'), 4)
    IDX_Q = next((i for i, h in enumerate(headers) if h == '试题题目'), 8)
    IDX_OPTS_START = next((i for i, h in enumerate(headers) if h and '选项A' in str(h)), 9)
    IDX_ANSWER = next((i for i, h in enumerate(headers) if h == '答案'), 17)
    IDX_EXPLAIN = next((i for i, h in enumerate(headers) if h == '题目解析'), 18)

    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) <= IDX_ID or not row[IDX_ID]: continue
        ans = row[IDX_ANSWER] if IDX_ANSWER < len(row) else None
        if not ans: continue
        ans_str = ans if isinstance(ans, str) else ''.join(ans)
        opts = []
        for j in range(8):
            if IDX_OPTS_START + j < len(row) and row[IDX_OPTS_START + j]:
                opts.append(row[IDX_OPTS_START + j])
        all_questions.append({
            'id': row[IDX_ID],
            'cat': row[IDX_CAT] if IDX_CAT < len(row) else '',
            'type': row[IDX_TYPE] if IDX_TYPE < len(row) else '单选题',
            'difficulty': row[IDX_DIFF] if IDX_DIFF < len(row) else '简单',
            'q': (row[IDX_Q] or '').replace('\r\n', ' ').replace('\n', ' ').strip(),
            'opts': opts,
            'answer': ans_str,
            'explain': (row[IDX_EXPLAIN] or '暂无解析').strip(),
        })

print(f'共读取 {len(all_questions)} 道题')
type_cnt = Counter(q['type'] for q in all_questions)
cat_cnt = Counter(q['cat'] for q in all_questions)
print(f'题型: {dict(type_cnt)}')
print(f'分类前5: {dict(list(cat_cnt.most_common(5)))}')

questions_json = json.dumps(all_questions, ensure_ascii=False)

# HTML 使用普通字符串（无 f-string，无反引号问题）
HTML = """<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1.0,maximum-scale=1.0">
<title>新型电力系统岗位题库 · 闯关版 v3</title>
<style>
* { margin: 0; padding: 0; box-sizing: border-box; -webkit-tap-highlight-color: transparent; }
body { font-family: -apple-system, "PingFang SC", "Microsoft YaHei", sans-serif; background: #f5f5f5; color: #333; max-width: 480px; margin: 0 auto; min-height: 100vh; padding-bottom: 20px; }
.header { background: white; padding: 14px 16px; box-shadow: 0 1px 3px rgba(0,0,0,0.06); position: sticky; top: 0; z-index: 100; display: flex; align-items: center; justify-content: space-between; }
.header h1 { font-size: 18px; font-weight: 600; flex: 1; text-align: center; }
.header .badge { background: linear-gradient(135deg, #f59e0b, #ef4444); color: white; padding: 4px 10px; border-radius: 12px; font-size: 11px; font-weight: 600; }
.progress-section { background: white; padding: 12px 16px; display: flex; align-items: center; gap: 12px; border-bottom: 1px solid #f0f0f0; position: sticky; top: 51px; z-index: 99; }
.progress-bar { flex: 1; height: 8px; background: #f3f4f6; border-radius: 4px; overflow: hidden; }
.progress-fill { height: 100%; background: linear-gradient(90deg, #f59e0b, #ef4444); transition: width 0.3s; }
.progress-text { font-size: 12px; color: #6b7280; font-weight: 600; }
.level-card { background: linear-gradient(135deg, #fff7ed, #fed7aa); margin: 12px 16px; padding: 16px; border-radius: 12px; border-left: 4px solid #f59e0b; }
.level-title { font-size: 14px; color: #9a3412; font-weight: 600; margin-bottom: 4px; }
.level-subtitle { font-size: 12px; color: #c2410c; }
.quiz-section { background: white; margin: 12px 16px; border-radius: 12px; padding: 18px; box-shadow: 0 1px 3px rgba(0,0,0,0.04); }
.q-header { display: flex; justify-content: space-between; align-items: center; margin-bottom: 14px; padding-bottom: 10px; border-bottom: 1px solid #f0f0f0; }
.q-meta { font-size: 12px; color: #999; }
.q-type { background: #fef3c7; color: #92400e; padding: 3px 10px; border-radius: 4px; font-size: 12px; font-weight: 600; }
.q-type.multi { background: #dbeafe; color: #1e40af; }
.q-text { font-size: 16px; line-height: 1.7; color: #1f2937; margin-bottom: 16px; }
.options { display: flex; flex-direction: column; gap: 10px; }
.option { display: flex; align-items: flex-start; gap: 10px; padding: 14px; border: 2px solid #e5e7eb; border-radius: 8px; cursor: pointer; transition: all 0.2s; background: #fafafa; }
.option:hover:not(.locked) { border-color: #f59e0b; background: #fffbeb; transform: translateX(2px); }
.option.selected { border-color: #f59e0b; background: #fffbeb; }
.option.locked { cursor: not-allowed; opacity: 0.85; }
.option.wrong { border-color: #ef4444; background: #fef2f2; color: #991b1b; animation: shake 0.4s; }
.option.correct-show { border-color: #10b981; background: #ecfdf5; color: #065f46; }
.opt-letter { width: 28px; height: 28px; border-radius: 50%; background: white; border: 2px solid #d1d5db; display: flex; align-items: center; justify-content: center; font-size: 13px; font-weight: 700; flex-shrink: 0; }
.option.selected .opt-letter { background: #f59e0b; color: white; border-color: #f59e0b; }
.option.correct-show .opt-letter { background: #10b981; color: white; border-color: #10b981; }
.opt-text { flex: 1; font-size: 14px; line-height: 1.5; }
@keyframes shake { 0%, 100% { transform: translateX(0); } 25% { transform: translateX(-6px); } 75% { transform: translateX(6px); } }
.feedback { margin-top: 14px; padding: 14px; border-radius: 8px; font-size: 14px; line-height: 1.6; display: none; }
.feedback.show { display: block; animation: fadeIn 0.3s; }
.feedback.correct { background: #ecfdf5; border-left: 4px solid #10b981; color: #065f46; }
.feedback b { display: block; margin-bottom: 4px; font-size: 15px; }
.feedback .icon { font-size: 22px; margin-right: 6px; vertical-align: middle; }
.feedback .reason { font-size: 13px; color: #4b5563; margin-top: 8px; padding-top: 8px; border-top: 1px solid rgba(0,0,0,0.06); }
@keyframes fadeIn { from { opacity: 0; transform: translateY(-4px); } to { opacity: 1; transform: translateY(0); } }
.actions { display: flex; gap: 8px; margin-top: 16px; }
.actions button { flex: 1; padding: 14px; border: none; border-radius: 8px; font-size: 15px; font-weight: 700; cursor: pointer; transition: all 0.2s; }
.btn-submit { background: #f59e0b; color: white; }
.btn-submit:hover:not(:disabled) { background: #d97706; }
.btn-submit:disabled { background: #d1d5db; color: #9ca3af; cursor: not-allowed; }
.btn-next { background: #10b981; color: white; animation: pulse 2s infinite; }
.btn-next:disabled { background: #d1d5db; color: #9ca3af; cursor: not-allowed; animation: none; }
@keyframes pulse { 0% { box-shadow: 0 0 0 0 rgba(16, 185, 129, 0.5); } 70% { box-shadow: 0 0 0 10px rgba(16, 185, 129, 0); } 100% { box-shadow: 0 0 0 0 rgba(16, 185, 129, 0); } }
.fail-modal { position: fixed; top: 0; left: 0; right: 0; bottom: 0; background: rgba(0,0,0,0.6); display: none; align-items: center; justify-content: center; z-index: 200; }
.fail-modal.show { display: flex; animation: fadeIn 0.3s; }
.fail-box { background: white; padding: 28px; border-radius: 16px; max-width: 360px; margin: 0 20px; text-align: center; }
.fail-icon { font-size: 60px; margin-bottom: 12px; }
.fail-title { font-size: 22px; font-weight: 700; color: #ef4444; margin-bottom: 8px; }
.fail-text { font-size: 14px; color: #4b5563; margin-bottom: 20px; line-height: 1.6; white-space: pre-line; text-align: left; }
.fail-buttons { display: flex; gap: 8px; }
.fail-buttons button { flex: 1; padding: 12px; border: none; border-radius: 8px; font-size: 14px; font-weight: 600; cursor: pointer; }
.btn-back { background: #f59e0b; color: white; }
.btn-restart { background: #6b7280; color: white; }
.win-modal { position: fixed; top: 0; left: 0; right: 0; bottom: 0; background: linear-gradient(135deg, rgba(251,191,36,0.95), rgba(239,68,68,0.95)); display: none; align-items: center; justify-content: center; z-index: 200; }
.win-modal.show { display: flex; animation: fadeIn 0.5s; }
.win-box { background: white; padding: 32px; border-radius: 20px; max-width: 360px; margin: 0 20px; text-align: center; }
.win-icon { font-size: 80px; margin-bottom: 16px; animation: bounce 1s infinite; }
@keyframes bounce { 0%, 100% { transform: translateY(0); } 50% { transform: translateY(-10px); } }
.win-title { font-size: 26px; font-weight: 800; background: linear-gradient(135deg, #f59e0b, #ef4444); -webkit-background-clip: text; -webkit-text-fill-color: transparent; margin-bottom: 8px; }
.win-text { font-size: 14px; color: #4b5563; margin-bottom: 20px; line-height: 1.6; }
.win-stat { background: #f9fafb; padding: 14px; border-radius: 8px; margin-bottom: 20px; }
.win-stat div { font-size: 13px; color: #6b7280; margin: 4px 0; }
.win-stat div b { color: #1f2937; font-size: 18px; }
.win-box button { background: linear-gradient(135deg, #f59e0b, #ef4444); color: white; border: none; padding: 14px 28px; border-radius: 8px; font-size: 15px; font-weight: 700; cursor: pointer; }
.watermark { position: fixed; top: 50%; left: 50%; transform: translate(-50%, -50%) rotate(-30deg); font-size: 80px; color: rgba(0,0,0,0.04); font-weight: 900; pointer-events: none; z-index: 1; }
.privacy-badge { background: #fef3c7; border: 1px solid #fcd34d; color: #92400e; padding: 8px 12px; margin: 12px 16px; border-radius: 6px; font-size: 12px; line-height: 1.5; }
.update-badge { background: #dbeafe; border: 1px solid #93c5fd; color: #1e40af; padding: 8px 12px; margin: 12px 16px; border-radius: 6px; font-size: 12px; line-height: 1.5; }
</style>
</head>
<body>
<div class="watermark">内部资料</div>
<div class="header">
<h1>闯关答题 · 第 <span id="level-num">1</span> 关</h1>
<span class="badge">闯关版v3</span>
</div>
<div class="progress-section">
<span class="progress-text">已闯 <span id="passed">0</span> / <span id="total">__TOTAL__</span> 关</span>
<div class="progress-bar"><div class="progress-fill" id="progress-fill" style="width:0%"></div></div>
<span class="progress-text" id="streak-text">连胜 0</span>
</div>
<div class="privacy-badge">
内部资料 · 仅限本人学习使用<br>
闯关模式：必须选对才能进入下一关，做错<b>只回到上一关</b>（不是从头开始）。
</div>
<div class="update-badge">
题库已更新 · 2026-08-25<br>
共 <b>__TOTAL__ 题</b> · 来源：<b>公司通用性制度题库</b>（xlsx 更新版）
</div>
<div class="level-card">
<div class="level-title">当前关卡</div>
<div class="level-subtitle" id="level-desc">单选题 · 来自《公司通用性制度题库》</div>
</div>
<div id="quiz-container"></div>

<div class="fail-modal" id="fail-modal">
<div class="fail-box">
<div class="fail-icon">❌</div>
<div class="fail-title">闯关失败</div>
<div class="fail-text" id="fail-text">很遗憾，你答错了这一题。正确答案已显示。再接再厉！</div>
<div class="fail-buttons">
<button class="btn-back" id="btn-back" onclick="goBack()">回到上一关</button>
<button class="btn-restart" onclick="restartFromBeginning()">从头开始</button>
</div>
</div>
</div>

<div class="win-modal" id="win-modal">
<div class="win-box">
<div class="win-icon">🏆</div>
<div class="win-title">全部通关！</div>
<div class="win-text">恭喜你完成了全部题库的挑战！</div>
<div class="win-stat">
<div>已闯关数：<b><span id="win-passed">0</span></b> 关</div>
<div>最佳连胜：<b><span id="win-streak">0</span></b> 关</div>
<div>完成时间：<b><span id="win-time">0</span></b></div>
</div>
<button onclick="restartFromBeginning()">再来一次</button>
</div>
</div>

<script>
const QUESTIONS = __QUESTIONS__;
let current = 0;
let passed = 0;
let streak = 0;
let maxStreak = 0;
let locked = false;
let selectedIdxs = [];
let startTime = Date.now();
let lastPassed = 0;

document.getElementById('total').textContent = QUESTIONS.length;

function renderQuestion() {
  const q = QUESTIONS[current];
  const isMulti = q.type === '多选题' || q.type === '不定项选择题';
  locked = false;
  selectedIdxs = [];
  document.getElementById('level-num').textContent = current + 1;
  document.getElementById('level-desc').innerHTML = q.type + ' · 来自《' + q.cat + '》';
  document.getElementById('progress-fill').style.width = (current / QUESTIONS.length * 100) + '%';
  document.getElementById('passed').textContent = passed;
  document.getElementById('streak-text').textContent = '连胜 ' + streak;
  document.getElementById('quiz-container').innerHTML =
    '<div class="quiz-section">' +
    '<div class="q-header">' +
    '<span class="q-type ' + (isMulti ? 'multi' : '') + '">' + q.type + (isMulti ? '（多选）' : '') + '</span>' +
    '<span class="q-meta">📊 ' + q.difficulty + '</span>' +
    '</div>' +
    '<div class="q-text">' + q.q + '</div>' +
    '<div class="options" id="options">' +
    q.opts.map(function(opt, idx) {
      var letter = String.fromCharCode(65 + idx);
      return '<div class="option" onclick="selectOption(' + idx + ', ' + isMulti + ')" data-idx="' + idx + '">' +
             '<div class="opt-letter">' + letter + + '</div>' +
             '<div class="opt-text">' + opt + '</div>' +
             '</div>';
    }).join('') +
    '</div>' +
    '<div class="feedback" id="feedback"></div>' +
    '<div class="actions">' +
    (isMulti ? '<button class="btn-submit" id="btn-submit" onclick="submitMulti()" disabled>提交答案</button>' : '') +
    '<button class="btn-next" id="btn-next" onclick="nextQuestion()" disabled>' +
    (current === QUESTIONS.length - 1 ? '🏁 完成' : '下一关 →') +
    '</button>' +
    '</div>' +
    '</div>';
  window.scrollTo({top: 0, behavior: 'smooth'});
}

function selectOption(idx, isMulti) {
  if (locked) return;
  var opts = document.querySelectorAll('.option');
  if (!isMulti) {
    selectedIdxs = [idx];
    opts.forEach(function(o) { o.classList.remove('selected'); });
    opts[idx].classList.add('selected');
    judgeSingle();
  } else {
    if (selectedIdxs.includes(idx)) {
      selectedIdxs = selectedIdxs.filter(function(i) { return i !== idx; });
      opts[idx].classList.remove('selected');
    } else {
      selectedIdxs.push(idx);
      opts[idx].classList.add('selected');
    }
    document.getElementById('btn-submit').disabled = selectedIdxs.length === 0;
  }
}

function judgeSingle() {
  var q = QUESTIONS[current];
  var opts = document.querySelectorAll('.option');
  var correctIdxs = q.answer.split('').map(function(a) { return a.charCodeAt(0) - 65; });
  var isCorrect = selectedIdxs.length === correctIdxs.length && selectedIdxs.every(function(i) { return correctIdxs.includes(i); });
  if (isCorrect) handleCorrect(q, correctIdxs, opts);
  else handleWrong(q, correctIdxs, opts);
}

function submitMulti() {
  var q = QUESTIONS[current];
  var opts = document.querySelectorAll('.option');
  var correctIdxs = q.answer.split('').map(function(a) { return a.charCodeAt(0) - 65; });
  var correctSet = new Set(correctIdxs);
  var selectedSet = new Set(selectedIdxs);
  var isCorrect = correctSet.size === selectedSet.size && Array.from(correctSet).every(function(x) { return selectedSet.has(x); });
  if (isCorrect) handleCorrect(q, correctIdxs, opts);
  else handleWrong(q, correctIdxs, opts);
}

function handleCorrect(q, correctIdxs, opts) {
  locked = true;
  opts.forEach(function(o, idx) {
    o.classList.add('locked');
    if (correctIdxs.includes(idx)) o.classList.add('correct-show');
  });
  passed++;
  lastPassed = current;
  streak++;
  if (streak > maxStreak) maxStreak = streak;
  var fb = document.getElementById('feedback');
  fb.className = 'feedback show correct';
  fb.innerHTML = '<b><span class="icon">✅</span>闯关成功！答案：' + q.answer + '</b>' + (q.explain ? '<div class="reason">💡 ' + q.explain + '</div>' : '');
  var subBtn = document.getElementById('btn-submit');
  if (subBtn) subBtn.style.display = 'none';
  document.getElementById('btn-next').disabled = false;
  document.getElementById('streak-text').textContent = '连胜 ' + streak;
}

function handleWrong(q, correctIdxs, opts) {
  locked = true;
  opts.forEach(function(o, idx) {
    o.classList.add('locked');
    if (selectedIdxs.includes(idx) && !correctIdxs.includes(idx)) o.classList.add('wrong');
    if (correctIdxs.includes(idx)) o.classList.add('correct-show');
  });
  var correctSet = new Set(correctIdxs);
  var selectedSet = new Set(selectedIdxs);
  var missing = Array.from(correctSet).filter(function(x) { return !selectedSet.has(x); });
  var wrong = Array.from(selectedSet).filter(function(x) { return !correctSet.has(x); });
  var reason = '';
  if (missing.length > 0 && wrong.length === 0) {
    reason = '❌ 少选了 ' + missing.length + ' 个选项 (' + missing.map(function(i) { return String.fromCharCode(65 + i); }).join('/') + ')';
  } else if (wrong.length > 0 && missing.length === 0) {
    reason = '❌ 选错了 ' + wrong.length + ' 个选项 (' + wrong.map(function(i) { return String.fromCharCode(65 + i); }).join('/') + ')';
  } else if (wrong.length > 0 && missing.length > 0) {
    reason = '❌ 选错了 ' + wrong.length + ' 个 (' + wrong.map(function(i) { return String.fromCharCode(65 + i); }).join('/') + ')，少选了 ' + missing.length + ' 个 (' + missing.map(function(i) { return String.fromCharCode(65 + i); }).join('/') + ')';
  } else {
    reason = '❌ 答错了';
  }
  streak = 0;
  var targetLevel = current === 0 ? 1 : current;
  document.getElementById('fail-text').innerHTML = reason + '<br><br>正确答案：<b>' + q.answer + '</b>' + (q.explain ? '<br><br>💡 解析：' + q.explain : '') + '<br><br>📍 闯关失败将<b>回到第 ' + targetLevel + ' 关</b>重新挑战';
  var btnBack = document.getElementById('btn-back');
  btnBack.textContent = current === 0 ? '🔄 重新挑战第 1 关' : '↩️ 回到第 ' + current + ' 关';
  setTimeout(function() {
    document.getElementById('fail-modal').classList.add('show');
  }, 800);
  document.getElementById('streak-text').textContent = '连胜 0';
}

function nextQuestion() {
  if (current < QUESTIONS.length - 1) {
    current++;
    renderQuestion();
  } else {
    showWin();
  }
}

function goBack() {
  if (current > 0) current = current - 1;
  document.getElementById('fail-modal').classList.remove('show');
  renderQuestion();
}

function restartFromBeginning() {
  current = 0;
  passed = 0;
  streak = 0;
  maxStreak = 0;
  startTime = Date.now();
  document.getElementById('fail-modal').classList.remove('show');
  document.getElementById('win-modal').classList.remove('show');
  renderQuestion();
}

function showWin() {
  document.getElementById('win-passed').textContent = passed;
  document.getElementById('win-streak').textContent = maxStreak;
  var sec = Math.round((Date.now() - startTime) / 1000);
  var m = Math.floor(sec / 60);
  var s = sec % 60;
  document.getElementById('win-time').textContent = m + '分' + s + '秒';
  document.getElementById('win-modal').classList.add('show');
}

renderQuestion();
</script>
</body>
</html>
"""

html = HTML.replace('__TOTAL__', str(len(all_questions))).replace('__QUESTIONS__', questions_json)

OUT_HTML.parent.mkdir(parents=True, exist_ok=True)
OUT_HTML.write_text(html, encoding='utf-8')
print(f'\nv3 闯关版已生成: {OUT_HTML}')
print(f'   大小: {OUT_HTML.stat().st_size / 1024:.1f} KB')
print(f'   内嵌题数: {len(all_questions)}')
print(f'   题型: {dict(type_cnt)}')